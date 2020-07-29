from flask import flash, redirect, render_template, url_for,request,session,abort
from flask_login import login_required, login_user, logout_user,current_user
from openpyxl import load_workbook
from . import mods
from .. import db
from ..api import *
from ..models import *
from ..func_ import *
from ..sms import sms_
import urllib.request
import simplejson as json
import config as conf

title_ = "OSFO"
empty_form = []

@mods.route('/orders') 
@login_required
def orders():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	company_ = Company.query.filter_by(id=company.id).first()
	return render_template("mods/company/orders.html",title=title_+" | Orders",year=year(),company=company_)

@mods.route('/order/flag/fulfilled/<int:order>/') 
@login_required
def fillOrder(order):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	order_ = Order.query.filter_by(id=order,company_id=company.id).first()
	if order_:
		# mail_send_res = sendOrderConfirmation(order,"fulfilled")
		confirmation_msg = f"Your order has been fulfilled. Order#{order_.id} has been confirmed [fulfilled] by #{order_.company.name}. You can now pickup your items and complete transaction @ store location."
		params = {"type":"fulfilled","order":order_.id,"message":confirmation_msg,"name":order_.customer.name,"app_link":url_for('osf.land',_external=True),"rec_link":url_for('osf.company',name=order_.company.name,_id=order_.company.id,_external=True),"email":order_.customer.email,"contact":Markup(f'Email: {conf.COMPANY_EMAIL}<br>Mobile: {conf.COMPANY_MOBILE}<br>Telephone: {conf.COMPANY_TELEPHONE}<br><span style="font-size:12px;">You can also reach out to us using the contact us form on the portal.</span><br>')}
		company.owner.add_notification("confirmation", "company", company.owner.id, "Order Fulfilled Confirmation", json.dumps(params, default=json_util.default), 0)
		if order_.customer.contact:
			sms_(order_.contact.contact, confirmation_msg)
		order_.status = 1
		error = db_commit_update_or_revert()
		flash("fulfilment notification will be sent to customer.")

	return redirect(url_for('mods.orders'))

@mods.route('/order/flag/paid/<int:order>/') 
@login_required
def paidOrder(order):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	order_ = Order.query.filter_by(id=order,company_id=company.id).first()
	if order_:
		order_.paid = 1
		order_.amount_due = 0
		error = db_commit_update_or_revert()
		if not error:
			# TODO send notification to customer of order fulfilment with pickup / delivery options
			flash("fulfilment alert has been sent to customer.")

		return redirect(url_for('mods.orders'))

@mods.route('/customers/')
@login_required
def customers():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	customers = getCustomersByCompany(company.id)
	return render_template("mods/customer/customers.html",title=title_+" | Customers",year=year(),customers=customers,company=company)

@mods.route('/customer/add/')
@login_required
def customerAdder():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	customers = getCustomersByCompany(company.id)
	return render_template("mods/customer/customer_adder.html",title=title_+" | Customers",form='add',year=year(),customers=customers,customer=[1],company=company)

@mods.route('/customers/create', methods=['GET','POST'])
@login_required
def addCustomer():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	ajax = True if "ajax" in form else False
		
	if ajax:
		res = createCustomer(form,"")
		customers = getCustomersByCompany(session["company"])
		
		ret = {}
		ret["result"] = {}
		ret["result"]["form"] = [res["form"]] if res["form"] else empty_form
		ret["result"]["types"] = [customers]
		ret["result"]["error"] = res["error"] 

		return json.dumps(ret,default=lambda a: {'id':a.id,'name':a.name } )
	else:
		res = createCustomer(form,request.files["image_"])	
		customers = getCustomersByCompany(session["company"])
		form_action = 'add'
	
		msg, msg_type = flash_res_msg(res)
		if msg:
			flash(msg,msg_type)


		if "action" in res and res["action"] == 'update':
			form_action = 'edit'

		return render_template("mods/customer/customer_adder.html",title=title_+" | Customers",form=form_action,year=year(),customers=customers,customer=res["form"],company=company)


@mods.route('/customers/edit/<int:customer_id>/', methods=['GET','POST'])
@login_required
def editCustomer(customer_id):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	customers = getCustomersByCompany(company_id)
	customer = [1]
	if customer_id:
		customer = [getCustomerbyID(customer_id)]
	
	return render_template("mods/customer/customer_adder.html",title=title_+" | Customers",form='edit',year=year(),customers=customers,customer=customer,company=company)


@mods.route('/customers/remove/',methods=['GET','POST'])
@login_required
def removeCustomer():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = deleteCustomer(form)
	customers = getCustomersByCompany(company.id)

	msg, msg_type = flash_res_msg(res)
	if msg:
		flash(msg,msg_type)

	return render_template("mods/customer/customers.html",title=title_+" | Customers",year=year(),customers=customers,customer=res["form"],company=company)
	

@mods.route("/company/open/<int:company>/",methods=["POST","GET"])
@login_required
def openCompany(company):
	company_ = Company.query.filter_by(id=company).first()
	company_.closed = 0
	error = db_commit_update_or_revert()
	error = createPOSSession(company)
	if "success" in error:
		flash("Pos session is now active.","success")

	if not error:
		flash("Open Sign Erected successfully.","success")

	return redirect(url_for('home.dashboard'))

@mods.route("/company/publish/<int:company>/",methods=["POST","GET"])
@login_required
def publishCompany(company):
	company_ = Company.query.filter_by(id=company).first()
	product_ = Product.query.filter_by(company_id=company,status=1).first()
	if company_:
		if len(company_.products) > 0:
			if product_:	
				company_.published = 1
				error = db_commit_update_or_revert()
				if not error:
					flash("Your Company is now available to customers for browsing and sending in orders. Be prepared to fulfil those orders.","success")
					return redirect(url_for('mods.openCompany',company=company_.id))
						
				else:
					flash(f"{error}")
			else:
				flash("You have no active products to complete the publish action.")
		else:
			flash("Your company setup is incomplete. Ensure that you have products for customers to shop and your company location is set.","warning")
	else:
		abort(404)
	return redirect(url_for('home.dashboard'))

@mods.route("/company/retract/<int:company>/",methods=["POST","GET"])
@login_required
def retractCompany(company):
	company_ = Company.query.filter_by(id=company).first()
	company_.published = 0
	error = db_commit_update_or_revert()
	if not error:
		flash("Your Company has been retracted.")
	else:
		flash(f"{error}")

	return redirect(url_for('home.dashboard'))


@mods.route("/company/pause/<int:company>/",methods=["POST","GET"])
@login_required
def pauseCompany(company):
	company_ = Company.query.filter_by(id=company).first()
	company_.paused = 1
	error = db_commit_update_or_revert()
	if not error:
		flash("Your Company has been paused, which means comsumers can only browse your product listings.")

	return redirect(url_for('home.dashboard'))


@mods.route("/company/close/<int:company>/",methods=["POST","GET"])
@login_required
def closeCompany(company):
	company_ = Company.query.filter_by(id=company).first()
	vline_ = Vline.query.filter_by(company_id=company,created_date=today()).first()

	if vline_ and vline_.active:
		flash("Close virtual line before attempting to close business.")
	else:
		company_.closed = 1
		error = db_commit_update_or_revert()
		if not error:
			flash("Close Sign Erected successfully.")
		else:
			flash(error)
	return redirect(url_for('home.dashboard'))


@mods.route("/company/create/",methods=["POST","GET"])
@login_required
def addCompany():
	# company = Company.query.filter_by(id=session["company"]["id"]).first()
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	ajax = True if "ajax" in form else False
	
	if current_user:
		ctypes = getCompanyType()
		if "bus_name" in form:
			ohl_ = 1
			if form["bus_ohl"]:
				ohl_ = form["bus_ohl"]
				
			location = f'{form["bus_address"]},{form["bus_state"]},{form["bus_city"]},{form["bus_postal"]}'
			company_form = {"name":form["bus_name"],"twitter":form["bus_twitter"],"google":form["bus_google"],"coords":form["bus_coords"],"facebook":form["bus_facebook"],"contact":form["bus_phone"],"email":form["bus_email"],"tax":form["bus_tax"],"type_id":form["bus_type"],"order_hold_limit":ohl_,"location":location,"status":0,"update":1}
			#account to be verified b4 it becomes active
			res = createCompany(company_form,request.files["image_"],request.files["store_front_image"])
			msg, msg_type = flash_res_msg(res)
			if msg:
				flash(msg,msg_type)

			if "success" in res:
				company = res["company"]
				session["company"] = company.as_dict()

		else:
			location_arr = company.location.split(",")
			bus_address = location_arr[0]
			bus_state = None
			bus_city = None
			bus_postal = None
			if len(location_arr) > 1:
				bus_state = location_arr[1] 
			if len(location_arr) > 2:
				bus_city = location_arr[2]
			if len(location_arr) > 3:
				bus_postal = location_arr[3]

			company_form = {"bus_name":company.name,"bus_phone":company.contact,"bus_email":company.email,"bus_tax":company.tax,"bus_owner":current_user.fullname,"bus_type":company.company_type_id,"bus_address":bus_address,"bus_state":bus_state,"bus_city":bus_city,"bus_postal":bus_postal,"bus_facebook":company.facebook,"bus_twitter":company.twitter,"bus_instagram":company.instagram,"bus_coords":company.coords,"bus_ohl":company.order_hold_limit,"bus_google":company.google}#account to be verified b4 it becomes active

			form = company_form
		print(company)
		return render_template("mods/company/company.html",company_types=ctypes,form=form,company=company)
	return redirect(url_for("home.dashboard"))

@mods.route("/vline/startline/<int:company>/",methods=["POST","GET"])
@login_required
def startVline(company):
	vline_ = Vline.query.filter_by(company_id=company,created_date=today()).first()
	if vline_:
		if vline_.active:
			flash("Virtual line is already active.")
		else:
			vline_.active = 1
			vline_.serving = 1
			error = db_commit_update_or_revert()
			if not error:
				flash("Virtual line restarted")
	else:
		vline_ = Vline(company_id=company, created_date=today(), serving=0, active=1)
		error = db_commit_add_or_revert(vline_)
		if not error:
			flash("Virtual line started succesfully.")

	return redirect(url_for('home.dashboard'))


@mods.route("/vline/pauseline/<int:company>/",methods=["POST","GET"])
@login_required
def pauseVline(company):
	vline_ = Vline.query.filter_by(id=company,created_date=today()).first()
	if vline_.active:
		vline_.active = 0
		error = db_commit_update_or_revert()
		if not error:
			flash("Virtual line paused")

	return redirect(url_for('home.dashboard'))
	
@mods.route("/vline/closevline/<int:company>/",methods=["POST","GET"])
@login_required
def closeVline(company):
	vline_ = Vline.query.filter_by(company_id=company,created_date=today()).first()
	if vline_:
		vline_.serving = -1
		vline_.active = 0
		error = db_commit_update_or_revert()
		if not error:
			flash("Virtual line closed successfully.")
	else:
		flash("Virtual line does not exist.")

	return redirect(url_for('home.dashboard'))

@mods.route("/vline/vlineup/<int:company>/",methods=["POST","GET"])
@login_required
def vLineup(company):
	vline_ = Vline.query.filter_by(company_id=company,created_date=today()).first()
	company_ = Company.query.filter_by(id=company).first()
	return render_template("home/vlineup.html",title=title_+" | Customers",year=year(),vline_=vline_,company=company_)

@mods.route("/vline/serve/<int:company>/<int:line>/<int:user>/",methods=["POST","GET"])
@login_required
def serving(company,line,user):
	vlineup = Vlineup.query.filter_by(line_id=line,company_id=company,user_id=user).first()
	# vline_ = Vline.query.filter_by(company_id=company,created_date=today()).first()
	
	if vlineup:
		vlineup.vline.serving = vlineup.place_in_line
		error = db_commit_update_or_revert()
		if not error:
			flash(f"Serving #{vlineup.place_in_line}.","success")

	return redirect(url_for('mods.vLineup',company=company))

# [products routing]
@mods.route('/products/', methods=['GET', 'POST'])
@login_required
def products_():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	products = []
	res = {"form":[1]}
	products = userItems(current_user,"products",session["company"]["id"])
	return render_template("mods/product/products.html",title=title_+" | Products",user=current_user,products=products,product=res,company=company)
	

@mods.route('/product/update/', methods=['GET', 'POST'])
@login_required
def product_update():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	form = request.form
	error = 1
	if 'item' in form:
		product = Product.query.filter_by(item_code=form['item'],company_id=company.id).first()
		if product:
			if form['field'] == 'item_name':
				product.name = html_to_txt(form['value'])
			else:
				if form['field'] == 'item_qty':
					product.qty = html_to_txt(form['value'])
				else:
					if form['field'] == 'item_description':
						product.description = html_to_txt(form['value'])
					else:
						if form['field'] == 'item_price':
							product.price = html_to_txt(form['value'])
						else:
							if form['field'] == 'item_cost':
								product.cost = html_to_txt(form['value'])
							else:
								if form['field'] == 'item_taxable':
									product.taxable = html_to_txt(form['value'])
								else:
									if form['field'] == 'item_status':
										variant_ = Variation.query.filter_by(product_id=product.id).first()
										if variant_:
											product.status = 0
										else:
											product.status = bool(int(html_to_txt(form['value'])))

			error = db_commit_update_or_revert()
	
	if not error:
		error = 0

	return json.dumps(error)



@mods.route('/products/add/', methods=['GET','POST'])
@login_required
def productAdder():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	products = []

	user = current_user
	products = userItems(user,"products",company.id)
	categories = getCategoriesByCompany(company.id)
	uom = getUom()
	ptype = getProductType()

	res = [1]

	return render_template("mods/product/product_adder.html",title=title_+" | Products",user=user,year=year(),product=res,categories=categories,uoms=uom,types=ptype,company=company,form="add")

@mods.route('/products/create/', methods=['GET','POST'])
@login_required
def addProduct():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = createProduct(form,request.files["image_"])    
	product = [1]
	if "form" in res:
		product = res["form"]

	product_uom = form.getlist("product_uoms")
	product_category = form.getlist("product_categories")

	if "success" in res:
		product = res["product"]
		uom_list = form.getlist("product_uoms")
		category_list = form.getlist("product_categories")
		uoms_error = addProductUoms(product,uom_list)
		category_error = addProductCategories(product,category_list)
		flash(uoms_error,"error")
		flash(category_error,"error")
		product_category = getProductCategories(product)
		product_uom = getProductUom(product)

	categories = getCategoriesByCompany(company.id)
	uoms = getUom()
	
	form_action = 'add'
	msg, msg_type = flash_res_msg(res)
	if msg:
		flash(msg,msg_type)

	if 'action' in res and res['action'] == 'update':
		form_action = 'edit'
		form = DDOT(form)
		return redirect(url_for('mods.editProduct',product_id=form.item_code))
	
	return redirect(url_for('mods.productAdder'))

	# return render_template("mods/product/product_adder.html",title=title_+" | Products",form=form_action,year=year(),product=product,categories=categories,company=company,uoms=uoms,product_category=product_category,product_uom=product_uom)

@mods.route('/products/variant/<int:product>/', methods=['GET','POST'])
@login_required
def productVariant(product):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	product_ = Product.query.filter_by(id=product).first()
	return render_template("mods/product/product_variant.html",title=title_+" | Product | Variant",year=year(),company=company,product=product_)

@mods.route('/products/variant/add/', methods=['GET','POST'])
@login_required
def addProductVariant():
	entry = DDOT(request.form)
	if entry.product_id:
		product_ = Product.query.filter_by(id=entry.product_id).first()
		if product_:
			if "update" in entry and bool(int(entry.update)):
				variant_ = Variation.query.filter_by(id=entry.id).first()
			else:
				variant_ = Variation.query.filter_by(product_id=entry.product_id,name=entry.name).first()

			if not variant_:
				if entry.variant_image == "": 
					variant_image = request.files["image_"]
					image_path = save_uploaded_file(variant_image, conf.PRODUCT_IMAGES_DIR)
				else: 
					image_path = entry.variant_image

				new_variant = Variation(product_id=entry.product_id,name=entry.name,image=image_path,created_date=now(),price=entry.price,qty=entry.qty)
				error = db_commit_add_or_revert(new_variant)
				if not error:
					if bool(int(entry.bulk)):
						variant_product = Product.query.filter_by(item_code=entry.variant_code).first()
						if variant_product:
							variant_product.status = 0
							error = db_commit_update_or_revert()
					flash("Product Variant added successfully","success") 
				else:
					flash(f'Error: {error}','error')
			else:
				res = updateVariant(entry,request.files["image_"])
				msg, msg_type = flash_res_msg(res)
				if msg:
					flash(msg,msg_type)

	return redirect(url_for('mods.productVariant',product=entry.product_id))


@mods.route('/products/variant/remove/', methods=['GET','POST'])
@login_required
def removeProductVariant():
	product = DDOT(request.form)
	deleteVariant(request.form)
				
	return redirect(url_for('mods.productVariant', product=product.product_id))
	


@mods.route('/products/edit/<product_id>/', methods=['GET','POST'])
@login_required
def editProduct(product_id):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	product = [getProductByID(product_id,company)]
	product_categories = getCategoriesByCompany(company.id)

	return render_template("mods/product/product_adder.html",title=title_+" | Products",form='edit',year=year(),product=product,categories=product_categories,company=company)

@mods.route('/products/type', methods=['GET','POST'])
@login_required
def ProductType():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	ptype = getProductType()
	return render_template("mods/product/product_type.html",title=title_+" | Product Types",user=current_user,year=year(),types=ptype,type=[],company=company)
	

@mods.route('/products/type/create', methods=['GET','POST'])
@login_required
def addProductType():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = createProductType(form) 
	ptype = getProductType() 
	
	return render_template("mods/product/product_type.html",title=title_+" | Product Types",user=current_user,year=year(),types=ptype,type=res,company=company)
	

@mods.route('/products/uom', methods=['GET','POST'])
@login_required
def ProductUom():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	uoms = getUom()
	
	return render_template("mods/product/product_uom.html",title=title_+" | Product Unit Of Measure",user=current_user,year=year(),uoms=uoms,uom=[],company=company)
	
@mods.route('/products/uom/create', methods=['GET','POST'])
@login_required
def addProductUom():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name 
	form = request.form
	res = createUom(form) 
	uom = getUom()
	
	return render_template("mods/product/product_uom.html",title=title_+" | Product Unit Of Measure",user=current_user,year=year(),uoms=uom,uom=res,company=company)

@mods.route('/products/uom/remove', methods=['POST'])
@login_required
def removeUom():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	uoms = deleteUom(form)

	return render_template("mods/product/product_uom.html",title=title_+" | Products",user=current_user,year=year(),uoms=uoms,uom=form,company=company)

	
@mods.route('/products/remove', methods=['POST'])
@login_required
def removeProduct():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = deleteProduct(company.id,form)
	products = getProducts()

	return redirect(url_for('mods.products_'))

	# return render_template("mods/product/products.html",title=title_+" | Products",user=current_user,year=year(),products=products,product=res,company=company)
	


@mods.route('/products/category/remove/<int:category_id>', methods=['GET','POST'])
@login_required
def removeCategory(category_id):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	res = deleteCategory(company,category_id)
	categories = getCategoriesByCompany(company.id)
	msg, msg_type = flash_res_msg(res)
	if msg:
		flash(msg,msg_type)

	return redirect(url_for('mods.Category'))

@mods.route('/products/type/remove/', methods=['POST','GET'])
@login_required
def removeType():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = deleteProductType(form)
	ptype = getProductType()
	return render_template("mods/product/products_type.html",title=title_+" | Products Types",user=auth.Adminuser,year=year(),ptypes=ptype,type=res,company=company)

@mods.route('/products/company/category/add/<int:category_id>', methods=['GET','POST'])
def addCompanyCategory(category_id):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	entry = CompanyCategory.query.filter_by(company_id=company.id,category_id=category_id).first()
	if not entry:
		entry = CompanyCategory(company_id=company.id,category_id=category_id)
		error = db_commit_add_or_revert(entry)
		if error:
			flash(error)

	return redirect(url_for('mods.Category'))

@mods.route('/company/category/product/add/', methods=['GET','POST'])
def addCategoryProducts():
	form = DDOT(request.form)
	product_list = request.form.getlist("products_for_category")
	entries = []
	for item in product_list:
		entry = CategoryProduct.query.filter_by(category_id=form.category,product_id=item).first()
		if not entry:
			entry = CategoryProduct(category_id=form.category,product_id=item)
			entries.append(entry)
	error = db_insert(entries)
	return redirect(url_for('mods.Category'))

@mods.route('/company/category/product/remove/', methods=['GET','POST'])
def removeCategoryProducts():
	form = DDOT(request.form)
	product_list = request.form.getlist("products_for_category_removal")
	entries = 0
	for item in product_list:
		entry = CategoryProduct.query.filter_by(category_id=form.category,product_id=item).first()
		if entry:
			error = db.session.delete(entry)
			entries = entries + 1
	if entries:
		db.session.commit()
	return redirect(url_for('mods.Category'))



@mods.route('/products/category', methods=['GET','POST'])
@login_required
def Category():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name

	categories = getCategoriesByCompany(company.id)
	global_categories = getCategoriesByCompanyType(company.company_type_id)
	return render_template("mods/product/product_category.html",title=title_+" | Product Categories",year=year(),company_categories=categories,global_categories=global_categories,company=company)

@mods.route('/products/category/add/product/<int:product>/', methods=['GET','POST'])
@login_required
def addToCategory(product):
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	return render_template("mods/product/product_category.html",title=title_+" | Product Categories",user=current_user,year=year(),categories=category,company=company)

@mods.route('/products/category/create/', methods=['GET','POST'])
def addProductCategory():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	form = request.form
	res = createCategory(form,request.files["image_"],company) 
	categories = getCategoriesByCompany(company.id)
	msg, msg_type = flash_res_msg(res)
	if msg:
		flash(msg,msg_type)
	return redirect(url_for('mods.Category'))

	
@mods.route('/products/import/', methods=['POST','GET'])
@login_required
def bulkImporProduct():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	form = request.form
	res = deleteUom(form)
	uom = getUom() 


	return render_template("mods/product/import.html",title=title_+" | Product | Import",year=year(),company=company)

@mods.route('/products/import/bulk/', methods=['POST','GET'])
@login_required
def doProductImport():
	company = Company.query.filter_by(id=session["company"]["id"]).first()
	title_ = company.name
	import_file = request.files["import_file"]
	form = request.form
	workbook = load_workbook(import_file)
	spreadsheet = workbook.active
	required_fields = []
	products_to_import = []
	products_to_update = []
	field_map = {}
	AC_FORMAT = ['item id', 'item name','item description', 'item price', 'item inventory', 'item tags', 'item image', 'item taxable']
	for header_row in spreadsheet.iter_rows(min_row=1,max_row=1,min_col=1,values_only=True):
		header_row = list(map(lambda x:x.lower(),header_row))
		for field in AC_FORMAT:
			if field not in header_row:
				required_fields.append(field)
			else:
				field_map[field] = header_row.index(field)

	if len(required_fields) > 0:
		flash(f"Required fields are missing. [{','.join(required_fields)}]")
	else:
		for row in spreadsheet.iter_rows(min_row=2,min_col=1,values_only=True):
			item_path = row[field_map['item image']]
			if item_path:
				item_image = getItemFromPath(item_path)
			else:
				item_image = 'img/products/default.png'

			item_ = Product.query.filter_by(item_code=row[field_map['item id']]).first()
			if not item_:
				product_ = Product(item_code=row[field_map['item id']],created_date=now(),description=row[field_map['item description']], name=row[field_map['item name']], price=row[field_map['item price']], qty=row[field_map['item inventory']], taxable=row[field_map['item taxable']], status=1,image=item_image)
				if product_ not in products_to_import:
					products_to_import.append(product_)
			else:
				name = row[field_map['item name']]
				description = row[field_map['item description']]
				price = row[field_map['item price']]
				inventory = row[field_map['item inventory']]
				taxable = row[field_map['item taxable']]

				if name.upper() != 'KEEP':
					item_.name = name

				if description.upper() != "KEEP":
					item_.description = description

				if price.upper() != "KEEP":
					item_.price = price

				if inventory.upper() != "KEEP":
					item_.qty = inventory

				if taxable.upper() != "KEEP":
					item_.taxable = taxable

				products_to_update.append(item_)
	
		total_rows = len(products_to_update) + len(products_to_import)
	return render_template("mods/product/import.html",title=title_+" | Product | Import",form=form,year=year(),company=company,products_to_import=products_to_import,products_to_update=products_to_update,total_rows=total_rows,no_import_form=True)